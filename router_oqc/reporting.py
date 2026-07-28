from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import DeviceStatus


HEAD_FILL = PatternFill("solid", fgColor="D9EAF7")
PASS_FILL = PatternFill("solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")


def _fit_columns(ws) -> None:
    for col_idx, cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)


def save_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def save_excel(
    path: Path,
    status: DeviceStatus,
    metadata: dict[str, str],
    test_rows: list[dict[str, str]],
    execution_rows: list[dict[str, str]],
) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Device_Status"
    headers = [
        "Test Time", "Router IP", "Device Name", "Serial Number", "Firmware Version",
        "WiFi Driver Version", "Uptime", "CPU Usage", "Memory Usage", "LAN IP",
        "Subnet Mask", "DHCP Server", "LAN MAC", "Login Result", "Status Result",
        "Overall Result", "Fail Code",
    ]
    ws.append(headers)
    ws.append([
        metadata.get("test_time", ""),
        metadata.get("router_ip", ""),
        status.system.get("Device Name", ""),
        status.system.get("Serial Number", ""),
        status.system.get("Firmware Version", ""),
        status.system.get("WiFi Driver Version", ""),
        status.system.get("Uptime", ""),
        status.system.get("CPU Usage", ""),
        status.system.get("Memory Usage", ""),
        status.lan.get("IP Address", ""),
        status.lan.get("Subnet Mask", ""),
        status.lan.get("DHCP Server", ""),
        status.lan.get("MAC Address", ""),
        metadata.get("login_result", ""),
        metadata.get("status_result", ""),
        metadata.get("overall_result", ""),
        metadata.get("fail_code", ""),
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
    ws.freeze_panes = "A2"
    _fit_columns(ws)

    wan_ws = wb.create_sheet("WAN_Status")
    wan_headers = [
        "Test Time", "Serial Number", "Interface", "VLAN ID", "Connection Type",
        "Protocol", "IP Address", "Gateway", "Status",
    ]
    wan_ws.append(wan_headers)
    for item in status.wan:
        wan_ws.append([
            metadata.get("test_time", ""),
            status.system.get("Serial Number", ""),
            item.get("Interface", ""),
            item.get("VLAN ID", ""),
            item.get("Connection Type", ""),
            item.get("Protocol", ""),
            item.get("IP Address", ""),
            item.get("Gateway", ""),
            item.get("Status", ""),
        ])
    for cell in wan_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
    wan_ws.freeze_panes = "A2"
    _fit_columns(wan_ws)

    test_ws = wb.create_sheet("Test_Log")
    test_headers = ["Test ID", "Item", "Expected", "Actual", "Result", "Fail Code", "Duration ms"]
    test_ws.append(test_headers)
    for row in test_rows:
        test_ws.append([row.get(h, "") for h in test_headers])
    for cell in test_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
    for row in test_ws.iter_rows(min_row=2):
        result = row[4].value
        for cell in row:
            cell.fill = PASS_FILL if result == "PASS" else FAIL_FILL
    _fit_columns(test_ws)

    exe_ws = wb.create_sheet("Execution_Log")
    exe_headers = ["Time", "Step", "Result", "Message"]
    exe_ws.append(exe_headers)
    for row in execution_rows:
        exe_ws.append([row.get(h, "") for h in exe_headers])
    for cell in exe_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
    _fit_columns(exe_ws)

    wb.save(path)
